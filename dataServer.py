"""
LoRaWAN ChirpStack Flask 웹 서버

기능:
- ChirpStack HTTP Integration (포트 8088)으로부터 데이터 수신
- SQLite 데이터베이스에 저장
- 웹 인터페이스 (포트 80)로 최근 20개 데이터 표시
- 로그인 인증 (ID/Password)
- 10초마다 자동 새로고침
"""

from flask import Flask, request, jsonify, render_template, render_template_string, redirect, url_for, session, Response, abort, send_file
from datetime import datetime, timedelta
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill
import threading
import os
from functools import wraps
from urllib.parse import urlparse, parse_qs
from io import BytesIO

# Flask 앱 초기화
app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'  # 보안을 위해 변경 필요

# 설정
DB_PATH = 'seoultel015.db'
HTTP_INTEGRATION_PORT = 80
WEB_PORT = 80

# 로그인 정보 (실제 환경에서는 암호화된 데이터베이스 사용 권장)
USERS = {
    #'admin': 'admin123',
    'admin': 'admin015',
    #'user': 'password123'
}

# 데이터베이스 락
db_lock = threading.Lock()


# ==================== 데이터베이스 초기화 ====================

def init_database():
    """데이터베이스 및 테이블 초기화"""
    with db_lock:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # LoRaWAN 데이터 테이블
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS lorawan_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                device_name TEXT NOT NULL,
                dev_eui TEXT NOT NULL,
                temperature REAL,
                rssi INTEGER,
                snr REAL,
                f_port INTEGER,
                f_cnt INTEGER,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 인덱스 생성
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_timestamp 
            ON lorawan_data(timestamp)
        ''')
        
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_dev_eui 
            ON lorawan_data(dev_eui)
        ''')
        
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_created_at 
            ON lorawan_data(created_at)
        ''')
        
        conn.commit()
        conn.close()
        print(f"✓ 데이터베이스 초기화 완료: {DB_PATH}")


# 데이터베이스 초기화
init_database()
    

# ==================== 데이터베이스 함수 ====================

def save_lorawan_data(data):
    """LoRaWAN 데이터를 데이터베이스에 저장"""
    with db_lock:
        try:
            conn = sqlite3.connect(DB_PATH, timeout=30)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO lorawan_data 
                (timestamp, device_name, dev_eui, temperature, rssi, snr, f_port, f_cnt)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data.get('timestamp'),
                data.get('device_name'),
                data.get('dev_eui'),
                data.get('temperature'),
                data.get('rssi'),
                data.get('snr'),
                data.get('f_port'),
                data.get('f_cnt')
            ))
            
            conn.commit()
            record_id = cursor.lastrowid
            conn.close()
            
            return record_id
            
        except sqlite3.Error as e:
            print(f"✗ 데이터 저장 오류: {e}")
            return None


def get_latest_data(limit=20):
    """최근 데이터 조회"""
    with db_lock:
        try:
            conn = sqlite3.connect(DB_PATH, timeout=30)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT * FROM lorawan_data 
                ORDER BY created_at DESC 
                LIMIT ?
            ''', (limit,))
            
            columns = [description[0] for description in cursor.description]
            results = []
            
            for row in cursor.fetchall():
                results.append(dict(zip(columns, row)))
            
            conn.close()
            return results
            
        except sqlite3.Error as e:
            print(f"✗ 데이터 조회 오류: {e}")
            return []


def get_statistics():
    """통계 정보 조회"""
    with db_lock:
        try:
            conn = sqlite3.connect(DB_PATH, timeout=30)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT 
                    COUNT(*) as total_count,
                    COUNT(DISTINCT dev_eui) as device_count,
                    AVG(temperature) as avg_temp,
                    MIN(temperature) as min_temp,
                    MAX(temperature) as max_temp,
                    AVG(rssi) as avg_rssi
                FROM lorawan_data
            ''')
            
            row = cursor.fetchone()
            columns = [description[0] for description in cursor.description]
            
            conn.close()
            return dict(zip(columns, row))
            
        except sqlite3.Error as e:
            print(f"✗ 통계 조회 오류: {e}")
            return {}


# ==================== 인증 데코레이터 ====================

def login_required(f):
    """로그인 확인 데코레이터"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


# ==================== ChirpStack 데이터 처리 ====================

def process_chirpstack_data(payload):
    """ChirpStack HTTP Integration 데이터 처리"""
    try:
        # ChirpStack 데이터 구조 파싱
        device_info = payload.get('deviceInfo', {})
        data = payload.get('data', {})
        rx_info = payload.get('rxInfo', [{}])[0] if payload.get('rxInfo') else {}

        #print('deviceInfo:', device_info)
        #print('data:', data)
        #print('rxInfo:', rx_info)
        
        # 디코딩된 데이터 추출
        #object_data = data.get('object', {})
        object_data = payload.get('object', {})
        #print('object:', object_data)
        
        #print('timestamp', datetime.now().isoformat())
        #print('device_name', device_info.get('deviceName', 'Unknown'))
        #print('dev_eui', device_info.get('devEui', 'Unknown'))
        #print('temperature', object_data.get('temperature'))
        #print('rssi', rx_info.get('rssi'))
        #print('snr', rx_info.get('loRaSNR'))
        #print('f_port', data.get('fPort', 0))
        #print('f_cnt', data.get('fCnt', 0))

        #if 'data' in object_data:
            #temperature = object_data.get('temperature')
        #else:
            #temperature = 0

        #temperature = object_data.get('temperature', 0),
        #print('temperature:', temperature)
        #if temperature >= 128:
            #temperature -= 256
        #print('temperature:', temperature)

        # 처리된 데이터 생성
        processed_data = {
            'timestamp': datetime.now().isoformat(),
            'device_name': device_info.get('deviceName', 'Unknown'),
            'dev_eui': device_info.get('devEui', 'Unknown'),
            'temperature': object_data.get('temperature', 0),
            #'temperature': temperature,
            'rssi': rx_info.get('rssi', 0),
            #'snr': rx_info.get('loRaSNR'),
            'snr': rx_info.get('snr', 0),
            'f_port': data.get('fPort', 0),
            'f_cnt': data.get('fCnt', 0)
        }
        
        return processed_data
        
    except Exception as e:
        print(f"✗ ChirpStack 데이터 처리 오류: {e}")
        return None


# ==================== Flask 라우트 ====================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """로그인 페이지"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if username in USERS and password in USERS[username]:
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('index00'))
        else:
            #return render_template_string(LOGIN_TEMPLATE, error="잘못된 사용자명 또는 비밀번호입니다.")
            return render_template('login_template.html', error="잘못된 사용자명 또는 비밀번호입니다.")
    
    #return render_template_string(LOGIN_TEMPLATE)
    return render_template('login_template.html')


@app.route('/logout')
def logout():
    """로그아웃"""
    session.clear()
    return redirect(url_for('login'))


#@app.route('/')
#def index0():
    ##return render_template_string('')
    ##return Response("Bad Request", status=400)
    #abort(403, description="이 리소스에 접근할 권한이 없습니다")

@app.route('/')
@login_required
def index00():
    """메인 대시보드"""
    data = get_latest_data(20)
    stats = get_statistics()
    username = session.get('username', 'User')
    
    return render_template(
        'dashboard_template_20.html',
        data=data, 
        stats=stats, 
        username=username
    )


@app.route('/data10')
@login_required
def index10():
    """메인 대시보드"""
    data = get_latest_data(10)
    stats = get_statistics()
    username = session.get('username', 'User')
    
    return render_template(
        'dashboard_template_10.html',
        data=data, 
        stats=stats, 
        username=username
    )


@app.route('/data20')
@login_required
def index20():
    """메인 대시보드"""
    data = get_latest_data(20)
    stats = get_statistics()
    username = session.get('username', 'User')
    
    return render_template(
        'dashboard_template_20.html',
        data=data, 
        stats=stats, 
        username=username
    )


@app.route('/data50')
@login_required
def index50():
    """메인 대시보드"""
    data = get_latest_data(50)
    stats = get_statistics()
    username = session.get('username', 'User')
    
    return render_template(
        'dashboard_template_50.html',
        data=data, 
        stats=stats, 
        username=username
    )


@app.route('/data100')
@login_required
def index100():
    """메인 대시보드"""
    data = get_latest_data(100)
    stats = get_statistics()
    username = session.get('username', 'User')
    
    return render_template(
        'dashboard_template_100.html',
        data=data, 
        stats=stats, 
        username=username
    )


@app.route('/data1k')
@login_required
def index1k():
    """메인 대시보드"""
    data = get_latest_data(1000)
    stats = get_statistics()
    username = session.get('username', 'User')
    
    return render_template(
        'dashboard_template_1k.html',
        data=data, 
        stats=stats, 
        username=username
    )


@app.route('/data10k')
@login_required
def index10k():
    """메인 대시보드"""
    data = get_latest_data(10000)
    stats = get_statistics()
    username = session.get('username', 'User')
    
    return render_template(
        'dashboard_template_10k.html',
        data=data, 
        stats=stats, 
        username=username
    )

@app.route('/data30k')
@login_required
def index30k():
    """메인 대시보드"""
    data = get_latest_data(30000)
    stats = get_statistics()
    username = session.get('username', 'User')
    
    return render_template(
        'dashboard_template_30k.html',
        data=data, 
        stats=stats, 
        username=username
    )

@app.route('/api/data20')
@login_required
def api_data20():
    """API: 최근 데이터 JSON 형식으로 반환"""
    limit = request.args.get('limit', 20, type=int)
    data = get_latest_data(limit)
    return jsonify(data)

@app.route('/api/data50')
@login_required
def api_data50():
    """API: 최근 데이터 JSON 형식으로 반환"""
    limit = request.args.get('limit', 50, type=int)
    data = get_latest_data(limit)
    return jsonify(data)

@app.route('/api/data100')
@login_required
def api_data100():
    """API: 최근 데이터 JSON 형식으로 반환"""
    limit = request.args.get('limit', 100, type=int)
    data = get_latest_data(limit)
    return jsonify(data)

@app.route('/api/data10k')
@login_required
def api_data10k():
    """API: 최근 데이터 JSON 형식으로 반환"""
    limit = request.args.get('limit', 10000, type=int)
    data = get_latest_data(limit)
    return jsonify(data)

@app.route('/api/data30k')
@login_required
def api_data30k():
    """API: 최근 데이터 JSON 형식으로 반환"""
    limit = request.args.get('limit', 30000, type=int)
    data = get_latest_data(limit)
    return jsonify(data)


# SQLite 데이터를 엑셀로 변환하는 함수
def create_excel_from_db(query, params=None):
    with db_lock:
        try:
            # 데이터베이스 연결
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row  # 컬럼명 접근을 위해
            cursor = conn.cursor()
    
            # 쿼리 실행
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
    
            rows = cursor.fetchall()
    
            # 엑셀 워크북 생성
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "데이터"
            
            if rows:
                # 헤더 스타일 설정
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
                # 헤더 작성
                columns = rows[0].keys()
                for col_idx, column in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=column)
                    cell.font = header_font
                    cell.fill = header_fill
        
                # 데이터 작성
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, column in enumerate(columns, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row[column])
        
                # 열 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            conn.close()
    
            # 메모리에 엑셀 파일 저장
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
    
            return excel_file

        except sqlite3.Error as e:
            print(f"✗ 데이터 저장 오류: {e}")
            return None

@app.route('/download/all')
@login_required
def download_all():
    try:
        query = "SELECT * FROM lorawan_data"
        excel_file = create_excel_from_db(query)

        filename = f"seoul015_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/stats')
@login_required
def api_stats():
    """API: 통계 정보 JSON 형식으로 반환"""
    stats = get_statistics()
    return jsonify(stats)


#@app.route('/webhook', methods=['POST'])
@app.route('/uplink', methods=['POST'])
def chirpstack_webhook():
    """
    ChirpStack HTTP Integration Webhook
    포트 8088로 수신
    """
    query = parse_qs(urlparse(request.url).query)
    #print('query:', query)
    #if query['event'][0] in ['join']:
    if query['event'][0] not in ['up']:
        return jsonify({'not up packet': str(query)}), 500

    try:
        # JSON 데이터 수신
        payload = request.get_json()
        
        if not payload:
            return jsonify({'error': 'No data received'}), 400
        #else:
            #print('payload:', payload)
        
        # 데이터 처리
        object_data = payload.get('object', {})
        print('object_data:', object_data)

        device_info = payload.get('deviceInfo', {})
        device_name = device_info.get('deviceName', 'Unknown')
        dev_eui = device_info.get('devEui', 'Unknown')

        #data = payload.get('data', {})
        #rx_info = payload.get('rxInfo', [{}])[0] if payload.get('rxInfo') else {}
        rx_info = payload.get('rxInfo', [{}])[0]

        #print('deviceInfo:', device_info)
        #print('data:', data)
        #print('rxInfo:', rx_info)
        
        # 디코딩된 데이터 추출
        object_data = payload.get('object', {})
        #print('object:', object_data)
        
        #print('timestamp', datetime.now().isoformat())
        #print('device_name', device_info.get('deviceName', 'Unknown'))
        #print('dev_eui', device_info.get('devEui', 'Unknown'))
        #print('temperature', object_data.get('temperature'))
        #print('rssi', rx_info.get('rssi'))
        #print('snr', rx_info.get('loRaSNR'))
        #print('f_port', payload.get('fPort', 0))
        #print('f_cnt', payload.get('fCnt', 0))

        #if 'data' in object_data:
            #temperature = object_data.get('temperature')
        #else:
            #temperature = 0
        temperature = object_data.get('temperature', 0)
        print('temperature:', temperature)
        if temperature >= 128:
            temperature = temperature - 256
        #elif temperature == 0:
            #temperature = 0.1
        print('temperature:', temperature)

        # 처리된 데이터 생성
        #timestamp = datetime.now() + timedelta(hours=9)
        processed_data = {
            'timestamp': datetime.now().isoformat(),
            #'timestamp': timestamp.isoformat(),
            'device_name': device_info.get('deviceName', 'Unknown'),
            'dev_eui': device_info.get('devEui', 'Unknown'),
            #'temperature': object_data.get('temperature', 0),
            'temperature': temperature,
            'rssi': rx_info.get('rssi', 0),
            #'snr': rx_info.get('loRaSNR'),
            'snr': rx_info.get('snr', 0),
            'f_port': payload.get('fPort', 0),
            'f_cnt': payload.get('fCnt', 0)
        }
        
        #return processed_data

        #processed_data = process_chirpstack_data(payload)
        if processed_data:
            # 데이터베이스 저장
            record_id = save_lorawan_data(processed_data)
            
            print(f"✓ 데이터 수신 및 저장 (ID: {record_id})")
            print(f"  Device: {processed_data['device_name']} ({processed_data['dev_eui']})")
            print(f"  Temperature: {processed_data['temperature']}°C")
            print(f"  RSSI: {processed_data['rssi']} dBm")
            
            return jsonify({
                'status': 'success',
                'record_id': record_id,
                'message': 'Data received and stored'
            }), 200
        else:
            return jsonify({'error': 'Failed to process data'}), 400
            
    except Exception as e:
        print(f"✗ Webhook 오류: {e}")
        return jsonify({'error': str(e)}), 500


# ==================== 메인 실행 ====================

def run_http_integration_server():
    """HTTP Integration 서버 실행 (포트 8088)"""
    from werkzeug.serving import make_server
    
    server = make_server('0.0.0.0', HTTP_INTEGRATION_PORT, app)
    print(f"✓ ChirpStack HTTP Integration 서버 시작: http://0.0.0.0:{HTTP_INTEGRATION_PORT}")
    print(f"  Webhook URL: http://your-server-ip:{HTTP_INTEGRATION_PORT}/webhook")
    server.serve_forever()


if __name__ == '__main__':
    print("=" * 70)
    print("LoRaWAN ChirpStack Flask 웹 서버")
    print("=" * 70)
    
    # 데이터베이스 초기화
    #init_database()
    
    # 서버 시작
    print(f"\n웹 인터페이스: http://0.0.0.0:{WEB_PORT}")
    print(f"ChirpStack Webhook: http://0.0.0.0:{HTTP_INTEGRATION_PORT}/webhook")
    print("\n로그인 정보:")
    print("  사용자명: admin / 비밀번호: admin123")
    print("  사용자명: user / 비밀번호: password123")
    print("\n서버 시작 중...\n")
    
    # HTTP Integration 서버를 별도 스레드에서 실행
    #integration_thread = threading.Thread(target=run_http_integration_server, daemon=True)
    #integration_thread.start()
    
    # 웹 서버 실행 (포트 80)
    # 주의: 포트 80은 root 권한이 필요할 수 있습니다
    try:
        app.run(host='0.0.0.0', port=WEB_PORT, debug=False, threaded=True)
    except PermissionError:
        print("\n⚠️ 포트 80 사용 권한 오류")
        print("해결 방법:")
        print("  1. sudo로 실행: sudo python dataServer.py")
        print("  2. 또는 포트 변경: WEB_PORT = 8080")
        print("\n임시로 포트 8080에서 실행합니다...")
        app.run(host='0.0.0.0', port=80, debug=False, threaded=True)

